# -*- coding: utf-8 -*-
"""
controllers/dashboard_controller.py
FLUJO COMPLETO INTEGRADO
"""

import sys
import os
from typing import Optional, List, Dict
import csv
import traceback
from datetime import datetime

from PySide6.QtWidgets import QMessageBox, QFileDialog, QMainWindow, QScrollArea, QWidget, QDialog, QVBoxLayout, QLabel, QTextEdit, QDialogButtonBox
from PySide6.QtCore import QDate, Qt

current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

try:
    from db import auditoria as auditoria_db
except ImportError:
    auditoria_db = None

try:
    from db import boletas as boletas_db
except ImportError:
    boletas_db = None

try:
    from controllers.auditoria_controller import AuditoriaController
    from controllers.solicitudes_controller import SolicitudesController
    from controllers.cierre_controller import CierreController
    from controllers.boleta_controller import BoletaController
except ImportError as e:
    print(f"[ERROR] Error importando controladores: {e}")
    AuditoriaController = None
    SolicitudesController = None
    CierreController = None
    BoletaController = None


class JustificacionDialog(QDialog):
    def __init__(self, parent=None, titulo="Justificación", mensaje="Ingrese la justificación:"):
        super().__init__(parent)
        self.setWindowTitle(titulo)
        self.setMinimumSize(400, 200)
        layout = QVBoxLayout(self)
        
        lbl = QLabel(mensaje)
        lbl.setWordWrap(True)
        layout.addWidget(lbl)
        
        self.text_justificacion = QTextEdit()
        self.text_justificacion.setPlaceholderText("Escriba aquí la justificación...")
        layout.addWidget(self.text_justificacion)
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def get_justificacion(self):
        return self.text_justificacion.toPlainText()



from PySide6.QtCore import QDate, Qt, QTimer
class DashboardController:
    """Controlador principal del Dashboard"""
    
    def __init__(self, conexion, ui, usuario_data: dict):
        self.conexion = conexion
        self.ui = ui
        self.usuario = usuario_data
        
        print(f"[INFO] DashboardController inicializado")
        print(f"[INFO] Usuario: {usuario_data.get('nombre_completo', 'Unknown')}")
        print(f"[INFO] Rol: {usuario_data.get('rol', 'unknown')}")
        
        self.ui._controller_ref = self
        
        self._init_controllers()
        self._connect_signals()
        self._load_user_info()
        self._load_initial_data()
        
        self._open_windows = []

        # Timer 5 segundos
        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self._auto_refresh)
        self.refresh_timer.start(5000) 

        self.notif_timer = QTimer()
        self.notif_timer.timeout.connect(self.cargar_contador_notificaciones)
        self.notif_timer.start(10000)  
        
    def _auto_refresh(self):
        try:
            self.cargar_solicitudes()
            dia = self.ui.combo_cierre_dia.currentText()
            if dia:
                self._actualizar_vista_cierre(dia)
        except Exception as e:
            print(f"[ERROR] _auto_refresh: {e}")
    
    def _init_controllers(self):
        """Inicializa sub-controladores"""
        try:
            self.auditoria_ctrl = AuditoriaController(self.conexion, self.ui) if AuditoriaController else None
            print(f"[{'✓' if self.auditoria_ctrl else '✗'}] AuditoriaController")
        except Exception as e:
            print(f"[✗] AuditoriaController: {e}")
            self.auditoria_ctrl = None
        
        try:
            self.solicitudes_ctrl = SolicitudesController(
                self.conexion, 
                self.ui, 
                auditoria_ctrl=self.auditoria_ctrl
            ) if SolicitudesController else None
            print(f"[{'✓' if self.solicitudes_ctrl else '✗'}] SolicitudesController")
        except Exception as e:
            print(f"[✗] SolicitudesController: {e}")
            self.solicitudes_ctrl = None
        
        try:
            self.cierre_ctrl = CierreController(
                self.conexion,
                auditoria_controller=self.auditoria_ctrl,
                solicitud_controller=self.solicitudes_ctrl
            ) if CierreController else None
            print(f"[{'✓' if self.cierre_ctrl else '✗'}] CierreController")
        except Exception as e:
            print(f"[✗] CierreController: {e}")
            self.cierre_ctrl = None
    
    def _load_user_info(self):
        """Carga información del usuario"""
        try:
            nombre = self.usuario.get("nombre_completo") or self.usuario.get("nombre", "Usuario")
            extension = self.usuario.get("extension", "")
            rol = self.usuario.get("rol", "usuario")
            
            self.ui.actualizar_info_usuario(nombre, extension, rol)
            
            if self.auditoria_ctrl:
                self.auditoria_ctrl.registrar(extension, "login", f"Ingreso - Rol: {rol}")
            
            print(f"[✓] Info usuario cargada: {nombre} ({extension})")
        except Exception as e:
            print(f"[ERROR] _load_user_info: {e}")
    
    def _connect_signals(self):
        """Conecta señales UI"""
        try:
            self.ui.btn_actualizar.clicked.connect(self.cargar_solicitudes)
            self.ui.btn_nuevaSolicitud.clicked.connect(self.nueva_solicitud)
            self.ui.combo_boletas_dia.currentTextChanged.connect(self.cargar_boletas_por_dia)
            self.ui.btn_exportar_boletas.clicked.connect(self.exportar_boletas_csv)
            self.ui.btn_enviar_cierre_actual.clicked.connect(self.marcar_dia_listo)
            self.ui.btn_listo_dia.clicked.connect(self.enviar_cierre_actual)
            self.ui.btn_enviar_cierre_semanal.clicked.connect(self.enviar_cierre_semanal)
            self.ui.combo_cierre_dia.currentTextChanged.connect(self._on_cierre_dia_changed)
            self.ui.btn_aud_buscar.clicked.connect(lambda: self.buscar_auditoria(self.ui.input_aud_buscar.text()))
            self.ui.btn_aud_limpiar.clicked.connect(self.cargar_auditoria)
            self.ui.btn_exportarAuditoria.clicked.connect(self.exportar_auditoria_csv)
            self.ui.btn_aud_limpiar.clicked.connect(self.limpiar_auditoria)
            print("[✓] Señales conectadas")
        except Exception as e:
            print(f"[ERROR] _connect_signals: {e}")
    
    def _load_initial_data(self):
        """Carga datos iniciales"""
        try:
            self.cargar_solicitudes()
            dia = self.ui.combo_boletas_dia.currentText()
            if dia:
                self.cargar_boletas_por_dia(dia)
            self.cargar_auditoria()
            self.cargar_contador_notificaciones()  # ✅ AGREGAR ESTA LÍNEA
            print("[✓] Datos iniciales cargados")
        except Exception as e:
            print(f"[ERROR] _load_initial_data: {e}")
    
    # ==================== SOLICITUDES ====================
    
    def cargar_solicitudes(self):
        """Carga tabla con filtro por rol"""
        try:
            if not self.solicitudes_ctrl:
                return
            
            try:
                self.conexion.rollback()
            except:
                pass
            
            self.ui.table_solicitudes.setRowCount(0)
            rol = self.usuario.get("rol", "usuario")
            extension_usuario = self.usuario.get("extension", "")
            
            registros = self.solicitudes_ctrl.listar()
            
            # ✅ CORREGIDO: Filtrar por rol
            if rol == "usuario":
                # Usuario solo ve sus propias boletas
                registros = [r for r in registros if r.get("extension") == extension_usuario]
            elif rol == "encargado":
                # ✅ Encargado ve TODAS las boletas (como operador)
                print(f"[INFO] Encargado '{extension_usuario}' cargando todas las solicitudes")
                # No filtrar, ve todo
            elif rol == "operador":
                # Operador ve todas
                pass
            elif rol == "admin":
                # Admin ve todas
                pass
            
            # Contar usuarios únicos
            extensiones_unicas = set()
            for r in registros:
                ext = r.get("extension")
                if ext:
                    extensiones_unicas.add(ext)
            
            # Agregar filas
            for r in registros:
                self.ui.agregar_fila_solicitud(
                    r.get("numero_boleta", ""),
                    r.get("extension", ""),
                    r.get("nombre_usuario", ""),
                    str(r.get("total_copias", 0)),
                    rol_view=rol,
                    estado=r.get("estado", "Pendiente"),
                    extra_id=r.get("id")
                )
            
            self.ui.lbl_totalBoletas.setText(f"Total de boletas: {len(registros)}")
            self.ui.lbl_totalUsuarios.setText(f"Usuarios atendidos hoy: {len(extensiones_unicas)}")
            
            print(f"[✓] Cargadas {len(registros)} solicitudes para rol '{rol}'")
        except Exception as e:
            print(f"[ERROR] cargar_solicitudes: {e}")
            traceback.print_exc()
    
    def nueva_solicitud(self):
        """Abre formulario nueva boleta"""
        try:
            from ui.boleta_form_ui import Ui_BoletaForm
            
            window = QMainWindow()
            window.setWindowTitle("Boleta - Nueva Solicitud")
            window.setWindowFlags(Qt.Window)
            
            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)
            scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            
            content_widget = QWidget()
            ui_form = Ui_BoletaForm()
            ui_form.setupUi(content_widget)
            
            scroll_area.setWidget(content_widget)
            window.setCentralWidget(scroll_area)
            window.resize(800, 600)
            window.setFixedSize(800, 600)
            
            # ✅ Crear controller
            boleta_ctrl = BoletaController(
                ui=ui_form,
                conexion=self.conexion,
                usuario=self.usuario,
                auditoria_ctrl=self.auditoria_ctrl,
                solicitudes_ctrl=self.solicitudes_ctrl,
                modo="form",
                parent_window=window
            )
            
            window._boleta_controller = boleta_ctrl
            
            def on_close():
                self.cargar_solicitudes()
                if window in self._open_windows:
                    self._open_windows.remove(window)
            
            window.closeEvent = lambda event: (on_close(), event.accept())
            window.show()
            self._open_windows.append(window)
            
        except Exception as e:
            print(f"[ERROR] nueva_solicitud: {e}")
            traceback.print_exc()
            QMessageBox.critical(None, "Error", f"No se pudo abrir formulario:\n{e}")
    
    def ver_boleta(self, numero_boleta: str, boleta_id: int):
        """
        Abre boleta para ver/editar según rol:
        - Usuario: solo lectura
        - Operador: puede editar su parte (empaste, observaciones)
        - Admin: puede editar todo
        """
        try:
            from ui.boleta_form_ui import Ui_BoletaForm
            
            window = QMainWindow()
            window.setWindowTitle(f"Boleta {numero_boleta}")
            
            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)
            
            content_widget = QWidget()
            ui_form = Ui_BoletaForm()
            ui_form.setupUi(content_widget)
            
            scroll_area.setWidget(content_widget)
            window.setCentralWidget(scroll_area)
            window.resize(800, 600)
            
            # ✅ Modo según rol
            rol = self.usuario.get("rol", "usuario")
            if rol in ("operador", "admin"):
                modo = "edit"
            else:
                modo = "view"
            
            boleta_ctrl = BoletaController(
                ui=ui_form,
                conexion=self.conexion,
                usuario=self.usuario,
                auditoria_ctrl=self.auditoria_ctrl,
                solicitudes_ctrl=self.solicitudes_ctrl,
                modo=modo,
                parent_window=window,
                boleta_id=boleta_id
            )
            
            window._boleta_controller = boleta_ctrl
            
            def on_close():
                self.cargar_solicitudes()
                # Recargar cierre si cambió
                if rol in ("operador", "admin"):
                    dia = self.ui.combo_cierre_dia.currentText()
                    if dia:
                        self._actualizar_vista_cierre(dia)
                if window in self._open_windows:
                    self._open_windows.remove(window)
            
            window.closeEvent = lambda event: (on_close(), event.accept())
            window.show()
            self._open_windows.append(window)
            
        except Exception as e:
            print(f"[ERROR] ver_boleta: {e}")
            traceback.print_exc()
    
    def cambiar_estado_boleta(self, numero_boleta: str, extension: str, nuevo_estado: str):
        """Cambia estado y refresca vista"""
        try:
            if not self.solicitudes_ctrl:
                return
            
            operador = self.usuario.get("extension", "unknown")
            
            print(f"[INFO] Cambiando estado de {numero_boleta} a '{nuevo_estado}'")
            
            success = self.solicitudes_ctrl.actualizar_estado(
                numero_boleta, nuevo_estado, operador
            )
            
            if success:
                print(f"[✓] Estado actualizado exitosamente")
                
                # Refrescar tabla de solicitudes
                self.cargar_solicitudes()
                
                # ✅ Si es "Listo", refrescar vista de cierre
                if nuevo_estado.lower() == "listo":
                    print(f"[✓] Boleta marcada como LISTA, refrescando vista de cierre...")
                    dia = self.ui.combo_cierre_dia.currentText()
                    if dia:
                        self._actualizar_vista_cierre(dia)
                    print(f"[✓✓✓] Vista de cierre actualizada con boleta {numero_boleta}")
        except Exception as e:
            print(f"[ERROR] cambiar_estado_boleta: {e}")
            traceback.print_exc()
            
    def abrir_panel_notificaciones(self):
        """Panel de notificaciones con limpiar"""
        try:
            from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, 
                                        QTableWidgetItem, QHeaderView, QPushButton, QLabel)
            from PySide6.QtCore import Qt
            
            try:
                self.conexion.rollback()
            except:
                pass
            
            dialog = QDialog(None)
            dialog.setWindowTitle("🔔 Notificaciones")
            dialog.resize(800, 500)
            
            layout = QVBoxLayout(dialog)
            
            header = QLabel("Historial de notificaciones")
            header.setStyleSheet("font-size: 14px; font-weight: bold; padding: 10px;")
            layout.addWidget(header)
            
            table = QTableWidget()
            table.setColumnCount(4)
            table.setHorizontalHeaderLabels(["Boleta", "Tipo", "Mensaje", "Estado"])
            table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
            table.setSelectionBehavior(QTableWidget.SelectRows)
            table.setEditTriggers(QTableWidget.NoEditTriggers)
            
            extension = self.usuario.get("extension", "")
            
            def cargar_notificaciones():
                table.setRowCount(0)
                cur = self.conexion.cursor()
                cur.execute("""
                    SELECT id, numero_boleta, tipo, mensaje, visto
                    FROM notificaciones
                    WHERE extension = %s
                    ORDER BY created_at DESC
                    LIMIT 100
                """, (extension,))
                
                for row in cur.fetchall():
                    notif_id, boleta, tipo, mensaje, visto = row
                    r = table.rowCount()
                    table.insertRow(r)
                    
                    table.setItem(r, 0, QTableWidgetItem(boleta or ""))
                    table.setItem(r, 1, QTableWidgetItem(tipo))
                    
                    msg_item = QTableWidgetItem(mensaje[:50] + "..." if len(mensaje) > 50 else mensaje)
                    msg_item.setToolTip(mensaje)
                    table.setItem(r, 2, msg_item)
                    
                    estado_texto = "✓ Visto" if visto else "🔴 Nuevo"
                    estado_item = QTableWidgetItem(estado_texto)
                    if not visto:
                        estado_item.setBackground(Qt.yellow)
                    table.setItem(r, 3, estado_item)
                
                cur.close()
            
            cargar_notificaciones()
            layout.addWidget(table)
            
            # Marcar como vistas
            cur = self.conexion.cursor()
            cur.execute("""
                UPDATE notificaciones SET visto = TRUE
                WHERE extension = %s AND visto = FALSE
            """, (extension,))
            self.conexion.commit()
            cur.close()
            
            self.ui.lbl_notifCount.setText("0")
            self.ui.lbl_notifCount.setVisible(False)
            
            # Botones
            btn_layout = QHBoxLayout()
            
            btn_limpiar = QPushButton("🗑️ Limpiar todo")
            btn_limpiar.setStyleSheet("background-color: #E74C3C; color: white;")
            def limpiar_todo():
                confirm = QMessageBox.question(None, "Limpiar", 
                    "¿Eliminar todas las notificaciones?", QMessageBox.Yes | QMessageBox.No)
                if confirm == QMessageBox.Yes:
                    cur = self.conexion.cursor()
                    cur.execute("DELETE FROM notificaciones WHERE extension = %s", (extension,))
                    self.conexion.commit()
                    cur.close()
                    cargar_notificaciones()
                    QMessageBox.information(None, "Limpiado", "Notificaciones eliminadas")
            
            btn_limpiar.clicked.connect(limpiar_todo)
            
            btn_layout.addWidget(btn_limpiar)
            btn_layout.addStretch()
            
            btn_cerrar = QPushButton("Cerrar")
            btn_cerrar.clicked.connect(dialog.accept)
            btn_layout.addWidget(btn_cerrar)
            
            layout.addLayout(btn_layout)
            dialog.exec()
            
        except Exception as e:
            print(f"[ERROR] abrir_panel_notificaciones: {e}")
            traceback.print_exc()

    def cargar_contador_notificaciones(self):
        """Carga contador de notificaciones no vistas"""
        try:
            try:
                self.conexion.rollback()
            except:
                pass
            
            extension = self.usuario.get("extension", "")
            cur = self.conexion.cursor()
            cur.execute("""
                SELECT COUNT(*) FROM notificaciones
                WHERE extension = %s AND visto = FALSE
            """, (extension,))
            count = cur.fetchone()[0]
            cur.close()
            
            if count > 0:
                self.ui.lbl_notifCount.setText(str(count))
                self.ui.lbl_notifCount.setVisible(True)
            else:
                self.ui.lbl_notifCount.setVisible(False)
                
        except Exception as e:
            print(f"[ERROR] cargar_contador_notificaciones: {e}")

    def rechazar_boleta(self, numero_boleta: str, extension: str, nombre: str):
        """Rechaza boleta y notifica EN TIEMPO REAL"""
        try:
            if not self.solicitudes_ctrl:
                return
            
            dialog = JustificacionDialog(
                None, "Rechazar boleta",
                f"¿Por qué se rechaza la boleta {numero_boleta}?\n\n"
                "El usuario recibirá una notificación EN TIEMPO REAL."
            )
            
            if dialog.exec() != QDialog.Accepted:
                return
            
            comentario = dialog.get_justificacion()
            if not comentario.strip():
                QMessageBox.warning(None, "Comentario requerido",
                    "Debe proporcionar un comentario explicando el rechazo.")
                return
            
            print(f"[INFO] Rechazando boleta {numero_boleta} con comentario: {comentario}")
            
            success = self.solicitudes_ctrl.rechazar_boleta(numero_boleta, extension, comentario)
            
            if success:
                QMessageBox.information(None, "Boleta rechazada",
                    f"Boleta {numero_boleta} rechazada.\nSe notificó al usuario.")
                
                # ✅ REFRESH TABLA (esto eliminará la fila)
                print(f"[✓] Recargando tabla de solicitudes...")
                self.cargar_solicitudes()
            else:
                QMessageBox.critical(None, "Error", f"No se pudo rechazar {numero_boleta}")
        except Exception as e:
            print(f"[ERROR] rechazar_boleta: {e}")
            traceback.print_exc()
    
    
    def editar_solicitud(self):
        QMessageBox.information(None, "No implementado",
            "La función de editar solicitudes no está implementada.")
    
    # ==================== BOLETAS ====================
    
    def cargar_boletas_por_dia(self, dia: str):
        """Carga boletas del día"""
        try:
            if not self.solicitudes_ctrl or not dia:
                return
            
            rol = self.usuario.get("rol", "")
            
            # ✅ Encargado ve desde calendar_archive
            if rol == "encargado":
                try:
                    self.conexion.rollback()
                except:
                    pass
                
                cur = self.conexion.cursor()
                cur.execute("""
                    SELECT datos FROM calendar_archive
                    WHERE fecha >= CURRENT_DATE - INTERVAL '7 days'
                    ORDER BY fecha DESC
                """)
                
                metas = []
                for row in cur.fetchall():
                    datos = row[0]
                    if isinstance(datos, str):
                        import json
                        datos = json.loads(datos)
                    
                    metas.append({
                        "boleta": datos.get('numero_boleta', ''),
                        "ext": datos.get('extension', ''),
                        "nombre": datos.get('nombre_usuario', ''),
                        "paginas": datos.get('paginas', 0),
                        "estado": "Archivado",
                        "dia": datos.get('dia', '')
                    })
                
                cur.close()
                
                self.ui._boletas_data[dia] = metas
                self.ui._refresh_boletas_day_ui(dia)
                return
        except Exception as e:
            print(f"[ERROR] cargar_boletas_por_dia: {e}")
    
    def buscar_boletas(self, q: str):
        """Busca boletas"""
        try:
            if not self.solicitudes_ctrl:
                return
            
            q = (q or "").strip()
            if not q:
                self.cargar_boletas_por_dia(self.ui.combo_boletas_dia.currentText())
                return
            
            results = self.solicitudes_ctrl.buscar(q) or []
            
            for i in reversed(range(self.ui.boletas_day_layout.count())):
                item = self.ui.boletas_day_layout.itemAt(i)
                if item and item.widget():
                    w = item.widget()
                    self.ui.boletas_day_layout.removeWidget(w)
                    w.setParent(None)
            
            for r in results:
                meta = {
                    "boleta": r.get("numero_boleta", ""),
                    "ext": r.get("extension_creador", "") or r.get("extension", ""),
                    "nombre": r.get("nombre_creador", "") or r.get("nombre_usuario", ""),
                    "paginas": r.get("paginas", 0),
                    "estado": r.get("estado", "Pendiente"),
                    "dia": ""
                }
                card = self.ui._build_card_from_meta(meta, es_cierre=False)
                self.ui.boletas_day_layout.insertWidget(self.ui.boletas_day_layout.count()-1, card)
            
            print(f"[✓] {len(results)} resultados")
        except Exception as e:
            print(f"[ERROR] buscar_boletas: {e}")
    
    # ==================== AUDITORÍA ====================
    
    def cargar_auditoria(self):
        try:
            if not self.auditoria_ctrl:
                return
            
            self.ui.table_auditoria.setRowCount(0)
            registros = self.auditoria_ctrl.listar()
            
            for rec in registros:
                self.ui.agregar_auditoria(
                    rec.get("fecha", ""),
                    rec.get("extension", ""),
                    rec.get("accion", ""),
                    rec.get("detalle", "")
                )
            
            print(f"[✓] {len(registros)} registros auditoría")
        except Exception as e:
            print(f"[ERROR] cargar_auditoria: {e}")
    
    def buscar_auditoria(self, extension: str):
        try:
            if not self.auditoria_ctrl:
                return
            
            extension = (extension or "").strip()
            if not extension:
                self.cargar_auditoria()
                return
            
            registros = self.auditoria_ctrl.buscar_por_extension(extension)
            self.ui.table_auditoria.setRowCount(0)
            
            for rec in registros:
                self.ui.agregar_auditoria(
                    rec.get("fecha", ""),
                    rec.get("extension", ""),
                    rec.get("accion", ""),
                    rec.get("detalle", "")
                )
            
            print(f"[✓] {len(registros)} registros")
        except Exception as e:
            print(f"[ERROR] buscar_auditoria: {e}")
    
    # ==================== CIERRE ====================
    
    def _on_cierre_dia_changed(self, dia: str):
        """Handler cambio día cierre"""
        try:
            if not self.cierre_ctrl or not dia:
                return
            
            self._actualizar_vista_cierre(dia)
        except Exception as e:
            print(f"[ERROR] _on_cierre_dia_changed: {e}")
    
    def _actualizar_vista_cierre(self, dia: str):
        """Actualiza vista de cierre CON DATOS DE BD"""
        try:
            if not self.cierre_ctrl:
                return
            
            # ✅ Refrescar UI desde BD (no desde memoria)
            self.ui._refresh_cierre_day_ui_from_db(dia, self.conexion)
            
            # Obtener resumen y actualizar texto
            resumen = self.cierre_ctrl.obtener_resumen_dia(self.ui, dia)
            texto = self.cierre_ctrl.generate_cierre_text(resumen)
            self.ui.actualizar_texto_boleta_cierre(texto)
            
            print(f"[✓] Vista cierre actualizada: {dia}")
        except Exception as e:
            print(f"[ERROR] _actualizar_vista_cierre: {e}")
    
    def marcar_dia_listo(self):
        """Abre formulario de cierre"""
        try:
            if not self.cierre_ctrl:
                return
            
            dia = self.ui.combo_cierre_dia.currentText()
            self.cierre_ctrl.abrir_cierre_form(self.ui, dia)
        except Exception as e:
            print(f"[ERROR] marcar_dia_listo: {e}")
    
    def enviar_cierre_semanal(self):
        """Envía cierre semanal (solo viernes)"""
        try:
            if QDate.currentDate().dayOfWeek() != 5:
                QMessageBox.warning(None, "No permitido",
                    "El cierre semanal solo se envía los viernes")
                return
            
            confirm = QMessageBox.question(None, "Enviar cierre semanal",
                "¿Confirma envío del cierre semanal?", QMessageBox.Yes | QMessageBox.No)
            
            if confirm == QMessageBox.Yes:
                QMessageBox.information(None, "Cierre semanal", "Cierre enviado exitosamente")
        except Exception as e:
            print(f"[ERROR] enviar_cierre_semanal: {e}")
    
    def enviar_cierre_actual(self):
        """Marca día listo y envía a encargado"""
        try:
            if not self.cierre_ctrl:
                return
            
            dia = self.ui.combo_cierre_dia.currentText()
            
            dow = QDate.currentDate().dayOfWeek()
            dias_map = {1:"Lunes",2:"Martes",3:"Miércoles",4:"Jueves",5:"Viernes"}
            dia_sistema = dias_map.get(dow, "Viernes")
            
            if dia != dia_sistema:
                QMessageBox.warning(None, "Día incorrecto",
                    f"Solo se puede enviar cierre del día actual ({dia_sistema})")
                return
            
            confirm = QMessageBox.question(None, "Enviar cierre",
                f"¿Enviar cierre del día {dia} a encargado?",
                QMessageBox.Yes | QMessageBox.No)
            
            if confirm == QMessageBox.Yes:
                self.cierre_ctrl.enviar_cierre(destinatario_tipo="encargado")
        except Exception as e:
            print(f"[ERROR] enviar_cierre_actual: {e}")
    
    # ==================== EXPORTACIONES ====================
    
    def exportar_boletas_csv(self):
        try:
            path, _ = QFileDialog.getSaveFileName(None, "Exportar boletas", "", "CSV Files (*.csv)")
            if not path:
                return
            
            todas = []
            for dia, items in self.ui._boletas_data.items():
                for it in items:
                    todas.append([dia, it.get("boleta"), it.get("ext"),
                                 it.get("nombre"), it.get("paginas"), it.get("estado")])
            
            with open(path, "w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                writer.writerow(["día", "boleta", "extensión", "nombre", "páginas", "estado"])
                writer.writerows(todas)
            
            QMessageBox.information(None, "Exportar", f"Exportado: {len(todas)} boletas")
        except Exception as e:
            print(f"[ERROR] exportar_boletas_csv: {e}")
    
    def exportar_auditoria_csv(self):
        try:
            path, _ = QFileDialog.getSaveFileName(None, "Exportar auditoría", "", 
                "Excel Files (*.xlsx);;CSV Files (*.csv)")
            if not path:
                return
            
            if not self.auditoria_ctrl:
                return
            
            registros = self.auditoria_ctrl.listar()
            
            if path.endswith('.xlsx'):
                # Exportar a Excel con formato
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment, PatternFill
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Auditoría"
                    
                    # Encabezados
                    headers = ["Fecha", "Extensión", "Acción", "Detalle"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(1, col, header)
                        cell.font = Font(bold=True, size=12)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Datos
                    for row_idx, rec in enumerate(registros, 2):
                        ws.cell(row_idx, 1, rec.get("fecha", ""))
                        ws.cell(row_idx, 2, rec.get("extension", ""))
                        ws.cell(row_idx, 3, rec.get("accion", ""))
                        ws.cell(row_idx, 4, rec.get("detalle", ""))
                    
                    # Ajustar anchos
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 15
                    ws.column_dimensions['C'].width = 25
                    ws.column_dimensions['D'].width = 50
                    
                    wb.save(path)
                    QMessageBox.information(None, "Exportado", f"Exportado: {len(registros)} registros")
                except ImportError:
                    QMessageBox.warning(None, "Librería faltante", 
                        "Instalar openpyxl: pip install openpyxl")
            else:
                # CSV simple
                import csv
                with open(path, "w", newline="", encoding="utf-8") as fh:
                    writer = csv.writer(fh)
                    writer.writerow(["fecha", "extensión", "acción", "detalle"])
                    for r in registros:
                        writer.writerow([r.get("fecha"), r.get("extension"),
                                        r.get("accion"), r.get("detalle")])
                QMessageBox.information(None, "Exportado", f"Exportado: {len(registros)} registros")
        except Exception as e:
            print(f"[ERROR] exportar_auditoria_csv: {e}")

    def limpiar_auditoria(self):
        """Limpia registros de auditoría"""
        try:
            confirm = QMessageBox.question(None, "Limpiar auditoría",
                "¿Eliminar TODOS los registros de auditoría?\n\nEsta acción no se puede deshacer.",
                QMessageBox.Yes | QMessageBox.No)
            
            if confirm == QMessageBox.Yes:
                cur = self.conexion.cursor()
                cur.execute("DELETE FROM auditoria")
                self.conexion.commit()
                cur.close()
                
                self.cargar_auditoria()
                QMessageBox.information(None, "Limpiado", "Auditoría eliminada")
        except Exception as e:
            print(f"[ERROR] limpiar_auditoria: {e}")

def cargar_boletas_cierre_encargado(self):
    """Carga boletas cerradas para encargado"""
    try:
        if not self.conexion:
            return
        
        try:
            self.conexion.rollback()
        except:
            pass
        
        # Obtener boletas cerradas del día
        cur = self.conexion.cursor()
        cur.execute("""
            SELECT numero_boleta, extension, nombre_usuario, total_copias, dia, fecha_procesado
            FROM boletas
            WHERE cerrado = TRUE
            ORDER BY fecha_procesado DESC
            LIMIT 100
        """)
        
        # Limpiar vista actual
        for i in reversed(range(self.ui.boletas_day_layout.count())):
            item = self.ui.boletas_day_layout.itemAt(i)
            if item and item.widget():
                w = item.widget()
                self.ui.boletas_day_layout.removeWidget(w)
                w.setParent(None)
        
        # Agregar cards
        for row in cur.fetchall():
            meta = {
                "boleta": row[0],
                "ext": row[1],
                "nombre": row[2],
                "paginas": row[3],
                "estado": "Cerrado",
                "dia": row[4]
            }
            card = self.ui._build_card_from_meta(meta, es_cierre=False)
            self.ui.boletas_day_layout.insertWidget(
                self.ui.boletas_day_layout.count()-1, card
            )
        
        cur.close()
    except Exception as e:
        print(f"[ERROR] cargar_boletas_cierre_encargado: {e}")